#!/usr/bin/env python3
"""
Import the tree-structured Projects feature (beacon_v2.project_items) from an
accounting-style Excel export.

Reads the **Projects** worksheet. Each Project# becomes a ROOT project_item
(local_id = Project#); every row that ALSO has a Phase value becomes a PHASE
imported as a DIRECT CHILD of that root (flat — dotted phase IDs like "1.1" /
"10.2" stay flat phases with that literal local_id, per product direction).
Phase "0" (Overhead) is valid.

Mapping (columns are matched by HEADER name, so column order can differ):
  Project#                     → local_id (root) / Phase → local_id (child)
  Project Name/Tasks           → name (root name overridden by the matching
                                  invoice project name when one exists)
  ClientID                     → matched to a beacon_v2.clients row → client_id
  Project Manager ID           → matched to a beacon_v2.users row → manager_user_id
  Contract Type                → enum (Fixed/Hourly/HNTE→hourly_nte/Overhead/…)
  Contract Amount              → contract_amount
  Total Labor Cost             → total_labor_cost
  Total Expense Cost           → total_expense_cost
  [A]TotalBilled(Services)Paid → billed_services
  [B]TotalBilled(Expenses)Paid → billed_expenses
  [C]TotalBilled(Taxes)Paid    → billed_taxes
  [A+B+C]Total Billed/Paid     → total_billed_paid

Matching:
  * Clients + PMs match case-insensitively, then FUZZY (difflib) so typos /
    middle-initials resolve (e.g. "Stuart Steiler" → Stuart Seiler, "James A.
    Wilson" → James Wilson). Genuinely-new names are left UNLINKED and reported.
  * The root gets item_type='main' when it has child phases (a container — no
    time/expense), else 'standard'. Phases are 'standard'. status='active'.

Robustness:
  * Inserts row-by-row and CATCHES per-row errors (duplicate phase id 23505,
    contract roll-up 23514) — they're reported, never abort the run.
  * If a project's phase amounts sum to MORE than its root contract amount,
    the root is inserted with a NULL cap (so every phase still imports) and the
    discrepancy is flagged for manual reconciliation.

Requires migration 20260624130000 (project_items) applied first.

Usage:
    python3 scripts/import_projects_xlsx.py                 # dry-run (read-only)
    python3 scripts/import_projects_xlsx.py --apply         # insert
    python3 scripts/import_projects_xlsx.py --wipe --apply  # delete all project_items first, then insert
    PROJECTS_XLSX=/path/to/file.xlsx python3 scripts/import_projects_xlsx.py
"""
from __future__ import annotations

import difflib
import os
import sys

import openpyxl
import requests
from dotenv import load_dotenv

load_dotenv(dotenv_path=os.path.join(os.getcwd(), ".env"))
URL = os.environ["SUPABASE_URL"]
KEY = os.environ["SUPABASE_SERVICE_KEY"]
XLSX = os.environ.get("PROJECTS_XLSX", os.path.expanduser("~/Downloads/MSMMActiveProjects.xlsx"))
SHEET = os.environ.get("PROJECTS_SHEET", "Projects")

H_READ = {"apikey": KEY, "Authorization": f"Bearer {KEY}", "Accept-Profile": "beacon_v2"}
H_WRITE = {**H_READ, "Content-Profile": "beacon_v2", "Content-Type": "application/json"}

# Fixed / Hourly / HNTE / Overhead etc. → beacon_v2.project_item_contract_type_enum.
CONTRACT_TYPE_MAP = {
    "fixed": "fixed",
    "hourly": "hourly",
    "hnte": "hourly_nte",
    "hourly not to exceed": "hourly_nte",
    "overhead": "overhead",
    "percentage": "percentage",
    "recurring": "recurring",
    "cost + percentage": "cost_plus_percentage",
    "cost + recurring": "cost_plus_recurring",
    "recurring + hourly": "recurring_plus_hourly",
}

# Known identity fixes (normalized Excel value → normalized canonical name).
# Fuzzy matching catches typos; these handle nicknames / longer official names
# that fuzzy/containment would miss.
PM_OVERRIDES = {
    "stuart steiler": "stuart seiler",   # typo
    "james a wilson": "jim wilson",      # Jim is the Beacon display name for James
}
CLIENT_OVERRIDES = {
    "st tammany parish": "st tammany parish government",
}

FUZZY = 0.84  # difflib ratio threshold for a confident name match


def norm(s) -> str:
    return " ".join(str(s or "").strip().lower().replace(".", " ").split())


def numv(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


class Matcher:
    """Exact-then-fuzzy name → record matcher with an overrides hook."""

    def __init__(self, records, name_keys, overrides):
        self.records = records
        self.overrides = overrides
        self.index = {}          # normalized name → record
        for r in records:
            for k in name_keys:
                nm = norm(r.get(k))
                if nm:
                    self.index.setdefault(nm, r)
        self.names = list(self.index.keys())

    def match(self, raw):
        nm = norm(raw)
        if not nm:
            return None, None
        nm = self.overrides.get(nm, nm)
        if nm in self.index:
            return self.index[nm], "exact"
        hit = difflib.get_close_matches(nm, self.names, n=1, cutoff=FUZZY)
        if hit:
            return self.index[hit[0]], f"fuzzy:{round(difflib.SequenceMatcher(None, nm, hit[0]).ratio(), 2)}"
        return None, None


def col_index(header):
    """Map our logical fields to actual column positions by header text."""
    H = [norm(h) for h in header]

    def find(*needles):
        for i, h in enumerate(H):
            if any(n in h for n in needles):
                return i
        return None

    return {
        "project": find("project#", "project #"),
        "phase": find("phase"),
        "name": find("project name", "tasks"),
        "client": find("clientid", "client id", "client"),
        "pm": find("project manager", "manager id", "manager"),
        "ctype": find("contract type"),
        "camount": find("contract amount"),
        "labor": find("labor cost"),
        "expense": find("expense cost"),
        "bserv": find("(services)"),
        "bexp": find("(expenses)"),
        "btax": find("(taxes)"),
        "btotal": find("total billed/paid", "[a+b+c]"),
    }


def main():
    apply = "--apply" in sys.argv
    wipe = "--wipe" in sys.argv

    # --- reference data from beacon_v2 ---
    clients = requests.get(f"{URL}/rest/v1/clients", headers=H_READ,
                           params={"select": "id,name,district"}, timeout=30).json()
    users = requests.get(f"{URL}/rest/v1/users", headers=H_READ,
                         params={"select": "id,first_name,last_name,display_name"}, timeout=30).json()
    invoices = requests.get(f"{URL}/rest/v1/anticipated_invoice", headers=H_READ,
                            params={"select": "project_number,project_name"}, timeout=30).json()
    existing_roots = requests.get(f"{URL}/rest/v1/project_items", headers=H_READ,
                                  params={"select": "local_id", "parent_id": "is.null"}, timeout=30).json()
    if not isinstance(existing_roots, list):
        # table not created yet (migration 20260624130000 not applied) — fine for dry-run
        print("  note: project_items not queryable yet (apply migration 20260624130000 before --apply)\n")
        existing_roots = []

    for u in users:
        u["full"] = " ".join(filter(None, [u.get("first_name"), u.get("last_name")]))
    # USACE-style clients store the district separately ("USACE" / "MVN-New
    # Orleans District"), but the sheet writes "USACE MVN". Index extra keys so
    # name+district and name+district-code ("USACE MVN") both resolve.
    for c in clients:
        nm, d = (c.get("name") or ""), (c.get("district") or "")
        code = d.split("-")[0].split("(")[0].strip()
        c["_kfull"] = f"{nm} {d}".strip()
        c["_kcode"] = f"{nm} {code}".strip() if code else nm
    client_m = Matcher(clients, ["name", "_kfull", "_kcode"], CLIENT_OVERRIDES)
    user_m = Matcher(users, ["display_name", "full"], PM_OVERRIDES)
    invoice_name = {}
    for iv in invoices:
        n = (iv.get("project_number") or "").strip()
        if n and n not in invoice_name and iv.get("project_name"):
            invoice_name[n] = iv["project_name"]
    existing = {(r.get("local_id") or "").strip() for r in existing_roots}

    # --- read the sheet ---
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"No '{SHEET}' worksheet (sheets: {wb.sheetnames})")
    ws = wb[SHEET]
    rows = [r for r in ws.iter_rows(values_only=True)]
    header, data = rows[0], [r for r in rows[1:] if any(c is not None for c in r)]
    ci = col_index(header)
    if ci["project"] is None or ci["phase"] is None:
        sys.exit(f"Could not locate Project#/Phase columns. Headers: {header}")

    def cell(r, key):
        i = ci[key]
        return r[i] if (i is not None and i < len(r)) else None

    # group rows by Project#, preserving order
    groups = {}
    for r in data:
        p = cell(r, "project")
        if p is None or str(p).strip() == "":
            continue
        groups.setdefault(str(p).strip(), []).append(r)

    print(f"Excel: {XLSX}  ·  sheet '{SHEET}'  ·  {len(groups)} projects, {len(data)} rows")
    print(f"Mode: {'APPLY' if apply else 'DRY-RUN'}{'  (--wipe)' if wipe else ''}\n")

    if wipe and apply:
        d = requests.delete(f"{URL}/rest/v1/project_items", headers=H_WRITE,
                            params={"id": "not.is.null"}, timeout=60)
        if d.status_code not in (200, 204):
            sys.exit(f"Wipe failed (HTTP {d.status_code}): {d.text[:200]}")
        print(f"  wiped existing project_items → HTTP {d.status_code}\n")
        existing = set()

    unmatched_clients, unmatched_pms = set(), set()
    n_proj = n_phase = n_skip_existing = 0
    errors = []

    def insert(payload):
        """POST one project_item; return (row|None, error_str|None)."""
        resp = requests.post(f"{URL}/rest/v1/project_items",
                             headers={**H_WRITE, "Prefer": "return=representation"},
                             json=payload, timeout=30)
        if resp.status_code in (200, 201):
            body = resp.json()
            return (body[0] if isinstance(body, list) else body), None
        return None, f"HTTP {resp.status_code}: {resp.text[:200]}"

    def map_ctype(raw):
        if raw is None:
            return None
        return CONTRACT_TYPE_MAP.get(norm(raw))

    for pid, rs in groups.items():
        root_rows = [r for r in rs if cell(r, "phase") is None or str(cell(r, "phase")).strip() == ""]
        phase_rows = [r for r in rs if not (cell(r, "phase") is None or str(cell(r, "phase")).strip() == "")]
        root_row = root_rows[0] if root_rows else rs[0]
        extra_root = " ⚠ multiple root rows" if len(root_rows) > 1 else ""

        # client / PM (resolved from the root row; same across the project)
        craw, praw = cell(root_row, "client"), cell(root_row, "pm")
        client, cway = client_m.match(craw)
        user, pway = user_m.match(praw)
        if craw and not client:
            unmatched_clients.add(str(craw).strip())
        if praw and not user:
            unmatched_pms.add(str(praw).strip())

        # name: prefer the matching invoice project name
        sheet_name = (cell(root_row, "name") or pid)
        inv = invoice_name.get(pid)
        name = inv or str(sheet_name).strip()
        name_src = "invoice" if inv else "sheet"

        # cap reconciliation: if phases sum over the root amount, drop the cap.
        root_amt = numv(cell(root_row, "camount"))
        phase_sum = sum(numv(cell(r, "camount")) or 0 for r in phase_rows)
        cap_warn = ""
        if root_amt is not None and phase_sum > root_amt + 0.01:
            cap_warn = f" ⚠ phases ${phase_sum:,.0f} > root ${root_amt:,.0f} → root cap dropped"
            root_amt = None

        cl = (f"{client['name']}({cway})" if client else (f"NO MATCH:{craw}" if craw else "—"))
        pm = (f"{user['full']}({pway})" if user else (f"NO MATCH:{praw}" if praw else "—"))
        print(f"[{pid}] {name[:46]:46s} {name_src:7s} client={cl[:28]:28s} pm={pm[:24]:24s} "
              f"{len(phase_rows)} phases{extra_root}{cap_warn}")

        if pid in existing and not wipe:
            print(f"      ↳ already imported — skipped (use --wipe to replace)")
            n_skip_existing += 1
            continue

        def base_fields(r):
            d = {
                "client_id": client["id"] if client else None,
                "manager_user_id": user["id"] if user else None,
                "contract_type": map_ctype(cell(r, "ctype")),
                "total_labor_cost": numv(cell(r, "labor")),
                "total_expense_cost": numv(cell(r, "expense")),
                "billed_services": numv(cell(r, "bserv")),
                "billed_expenses": numv(cell(r, "bexp")),
                "billed_taxes": numv(cell(r, "btax")),
                "total_billed_paid": numv(cell(r, "btotal")),
                "status": "active",
            }
            return d

        root_payload = {
            "local_id": pid,
            "parent_id": None,
            "name": name,
            "item_type": "main" if phase_rows else "standard",
            "contract_amount": root_amt,
            **base_fields(root_row),
        }

        if not apply:
            n_proj += 1
            n_phase += len(phase_rows)
            continue

        root, err = insert(root_payload)
        if err:
            errors.append(f"[{pid}] root insert failed — {err}")
            continue
        n_proj += 1
        root_id = root["id"]

        for r in phase_rows:
            ph = str(cell(r, "phase")).strip()
            payload = {
                "local_id": ph,
                "parent_id": root_id,
                "name": str(cell(r, "name") or f"Phase {ph}").strip(),
                "item_type": "standard",
                "contract_amount": numv(cell(r, "camount")),
                **base_fields(r),
            }
            _, perr = insert(payload)
            if perr:
                errors.append(f"[{pid}] phase '{ph}' skipped — {perr}")
            else:
                n_phase += 1

    # --- summary ---
    print("\n" + "=" * 72)
    print(f"{'Imported' if apply else 'Would import'}: {n_proj} projects, {n_phase} phases"
          + (f"  ·  skipped {n_skip_existing} already-present projects" if n_skip_existing else ""))
    if unmatched_clients:
        print(f"\nUNMATCHED CLIENTS ({len(unmatched_clients)}) — imported UNLINKED, add to the directory + re-run:")
        for c in sorted(unmatched_clients):
            print(f"   · {c}")
    if unmatched_pms:
        print(f"\nUNMATCHED PROJECT MANAGERS ({len(unmatched_pms)}) — imported UNLINKED:")
        for p in sorted(unmatched_pms):
            print(f"   · {p}")
    if errors:
        print(f"\nROW ERRORS ({len(errors)}):")
        for e in errors:
            print(f"   · {e}")
    if not apply:
        print("\n(dry-run — re-run with --apply to write)")


if __name__ == "__main__":
    main()
