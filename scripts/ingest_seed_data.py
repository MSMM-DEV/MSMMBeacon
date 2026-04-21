#!/usr/bin/env python3
"""
MSMM Beacon — seed-data ingester.

Parses the CSVs + xlsx under Data/ and populates the `beacon` schema
via Supabase PostgREST (using SUPABASE_SERVICE_KEY from .env).

  python3 scripts/ingest_seed_data.py --dry-run    # parse only, no writes
  python3 scripts/ingest_seed_data.py              # parse + upload
  python3 scripts/ingest_seed_data.py --wipe       # DELETE all data first, then upload
"""
from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "Data"
load_dotenv(ROOT / ".env")

URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
KEY = os.environ.get("SUPABASE_SERVICE_KEY", "")
if not URL or not KEY:
    sys.exit("Missing SUPABASE_URL / SUPABASE_SERVICE_KEY in .env")

REST = f"{URL}/rest/v1"
H_BASE = {
    "apikey": KEY,
    "Authorization": f"Bearer {KEY}",
    "Accept-Profile": "beacon",
    "Content-Profile": "beacon",
}
H_JSON = {**H_BASE, "Content-Type": "application/json"}


# --------------------------------------------------------------------------
# HTTP helpers
# --------------------------------------------------------------------------
def _req(method: str, path: str, *, params=None, json=None, prefer=None, headers_extra=None):
    hdrs = dict(H_JSON if json is not None else H_BASE)
    if prefer:
        hdrs["Prefer"] = prefer
    if headers_extra:
        hdrs.update(headers_extra)
    r = requests.request(method, f"{REST}/{path}", params=params, json=json, headers=hdrs, timeout=60)
    if r.status_code >= 400:
        raise RuntimeError(f"{method} {path} → {r.status_code}: {r.text[:500]}")
    return r


def preflight():
    r = requests.get(f"{REST}/users", headers=H_BASE, params={"limit": 1}, timeout=30)
    if r.status_code in (404, 406) or (
        r.status_code >= 400 and "schema" in r.text.lower() and "beacon" in r.text.lower()
    ):
        sys.exit(
            "ERROR: PostgREST cannot see the 'beacon' schema.\n"
            "Open Supabase Dashboard → Settings → API → Exposed schemas → add 'beacon'.\n"
            f"(response: {r.status_code} {r.text[:200]})"
        )
    r.raise_for_status()
    rows = r.json()
    print(f"Preflight OK — beacon.users returns {len(rows)} row(s).")


# --------------------------------------------------------------------------
# Parsing helpers
# --------------------------------------------------------------------------
_MONEY_STRIP = re.compile(r"[\$,\s]")


def parse_money(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "" or s in ("-", "#REF!", "N/A"):
        return None
    cleaned = _MONEY_STRIP.sub("", s)
    if cleaned in ("", ".", "-", "-."):
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_date_str(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    # Could not parse — return None; caller stashes raw text elsewhere if needed
    return None


_SUB_CELL_RE = re.compile(r"^\s*\$?\s*([\d,]+(?:\.\d+)?)?\s*(?:\(([^)]+)\))?\s*$")


def parse_sub_cell(v):
    """'$90,000 (survey)' → (90000.0, 'survey'). '(structural)' → (None, 'structural')."""
    if v is None:
        return (None, None)
    s = str(v).strip()
    if not s or s == "$0.00":
        return (None, None) if not s else (0.0, None)
    m = _SUB_CELL_RE.match(s)
    if not m:
        return (None, s)
    amt = float(m.group(1).replace(",", "")) if m.group(1) else None
    disc = m.group(2).strip() if m.group(2) else None
    return (amt, disc)


def split_client_name(name: str | None):
    """'USACE-MVN-New Orleans District' → ('USACE', 'MVN-New Orleans District'). Others unchanged."""
    if not name:
        return (None, None)
    name = str(name).strip()
    if not name:
        return (None, None)
    if name.startswith("USACE-"):
        parts = name.split("-", 2)
        if len(parts) == 3:
            return ("USACE", f"{parts[1]}-{parts[2]}")
        if len(parts) == 2:
            return ("USACE", parts[1])
    return (name, None)


_CLIENT_ALIASES = {
    "National Parks": "National Park Service",
    "National Parks ": "National Park Service",
    "Orleans Parish Schoolboard": "Orleans Parish School Board",
    "St. Charles Parish/Corp": "St. Charles Parish",
    "USACE New Oleans District": "USACE-MVN-New Orleans District",
    "New Orleans District": "USACE-MVN-New Orleans District",
    "MVM": "USACE-MVM-Memphis District",
    "STP": "St. Tammany Parish Government",
}


def canonicalize_client(raw: str | None):
    """Returns (name, district) after alias + split."""
    if not raw:
        return (None, None)
    raw = str(raw).strip().rstrip(":").strip()
    if not raw:
        return (None, None)
    aliased = _CLIENT_ALIASES.get(raw, raw)
    return split_client_name(aliased)


def split_semi_list(v):
    if not v:
        return []
    return [p.strip() for p in str(v).split(";") if p.strip()]


# --------------------------------------------------------------------------
# Known corrections
# --------------------------------------------------------------------------
# 3 rows in 2026 Potential where Client column actually held the prime firm.
# Key is the raw project name; value = (client_name, client_district, org_type, prime_company).
POTENTIAL_2026_SUB_FIXES = {
    "SWB-LSLR- Sub to CDM": ("Sewerage & Water Board of New Orleans", None, "City", "CDM"),
    "Westside Creek Amendment": ("USACE", "San Antonio", "Federal", "MBI/HZ JV"),
    "LPB Task Order - CPRA": ("Coastal Protection and Restoration Authority", None, "State", "Stantec"),
}

# Low-probability 2026 rows with blank Client column — infer client from project name.
# Value is (client_name, client_district, org_type) or None if unclear.
POTENTIAL_BLANK_CLIENT_INFERENCES: dict[str, tuple | None] = {
    "USFWS - Bayou Sauvage - CEI Mark Schexnayder": ("USFWS", None, "Federal"),
    "Port of New Orleans (Chris Gilmore) New Role": ("The Port of New Orleans", None, "City"),
    "Terrebonne CDBG MSA": ("Terrebonne Parish Consolidated Government", None, "Parish"),
    "JP I&I Parish wide ": ("Jefferson Parish", None, "Parish"),
    "Plaquemines Parish LNG-GIS": ("Plaquemines Parish", None, "Parish"),
    "RTA Zero Emissions": ("Regional Transit Authority", None, "Regional"),
    "RTA Union Passenger Terminal": ("Regional Transit Authority", None, "Regional"),
    "Entergy Future Projects": ("Entergy", None, "Other"),
    # Unclear — leave NULL for manual backfill:
    "Southwest Coastal Non Structural": None,
    "Sun Granite": None,
    "Fort Cavazos (MHZ)": None,
    "MSY-CMAR_Contractors for Sewer/WWTP": None,
    "Port Arthur Flood Wall- Sub to AECOM": None,
}


# --------------------------------------------------------------------------
# Parsers per source file
# --------------------------------------------------------------------------
def _read_csv(path: Path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.reader(f))


def parse_potential(path: Path, year: int):
    rows = _read_csv(path)
    # Find header row
    hdr_idx = next(i for i, r in enumerate(rows) if r and r[0].strip() == "Project Name")
    headers = rows[hdr_idx]
    is_2026 = "Prime or Sub?" in headers

    current_prob = "High"
    out: list[dict] = []
    for r in rows[hdr_idx + 1:]:
        if not r:
            continue
        cell0 = (r[0] or "").strip()
        if not cell0:
            continue
        if cell0.startswith("Total Amount ("):
            m = re.match(r"Total Amount \((\w+) Probability\)", cell0)
            if m:
                seen = m.group(1).lower()
                if seen == "high":
                    current_prob = "Medium"
                elif seen == "medium":
                    current_prob = "Low"
                else:
                    current_prob = None
            continue

        def cell(i):
            return r[i] if i < len(r) else ""

        if is_2026:
            subs_raw = [cell(5), cell(6), cell(7), cell(8)]
            item = {
                "year": year,
                "project_name": cell0,
                "role_raw": (cell(1) or "").strip(),
                "client_raw": (cell(2) or "").strip(),
                "total_contract_amount": parse_money(cell(3)),
                "msmm_amount": parse_money(cell(4)),
                "subs_raw": subs_raw,
                "pm_raw": (cell(9) or "").strip(),
                "notes": (cell(10) or "").strip() or None,
                "next_action_date": parse_date_str(cell(11)),
                "next_action_note": (cell(12) or "").strip() or None,
                "project_number": (cell(13) or "").strip() or None,
                "probability": current_prob,
            }
        else:  # 2025 — narrower schema; user said MSMM is prime for everything.
            subs_raw = [cell(3)]  # single Subs column
            item = {
                "year": year,
                "project_name": cell0,
                "role_raw": "Prime",
                "client_raw": "",  # no client column in 2025
                "total_contract_amount": parse_money(cell(1)),
                "msmm_amount": parse_money(cell(2)),
                "subs_raw": subs_raw,
                "pm_raw": (cell(4) or "").strip(),
                "notes": (cell(5) or "").strip() or None,
                "next_action_date": parse_date_str(cell(6)),
                "next_action_note": None,
                "project_number": (cell(7) or "").strip() or None,
                "probability": current_prob,
            }
        # Normalize project_number like "2026XX" → None (placeholder)
        if item["project_number"] and re.fullmatch(r"\d{4}[Xx]+", item["project_number"]):
            item["project_number_placeholder"] = item["project_number"]
            item["project_number"] = None
        out.append(item)
    return out


def parse_awaiting(path: Path):
    wb = load_workbook(path, data_only=True)
    ws = wb["Proposals"]
    rows = list(ws.iter_rows(values_only=True))
    # Header row has 'Proposal Year'
    hdr_idx = next(i for i, r in enumerate(rows) if r and r[0] == "Proposal Year")
    current_org = None
    out: list[dict] = []
    for r in rows[hdr_idx + 1:]:
        if not r or not any(r):
            continue
        cell0 = r[0]
        if isinstance(cell0, str) and cell0.strip().startswith("Org Type :"):
            current_org = cell0.split(":", 1)[1].strip()
            continue
        # data row
        py, title, client, prime, sub, status, stage, details, pool, sub_date, cc_no, msmm_no, exp, cap, used, remain = (
            list(r) + [None] * 16
        )[:16]
        out.append({
            "proposal_year": int(py) if py not in (None, "") else None,
            "title": (title or "").strip() or None,
            "client_raw": (client or "").strip() if isinstance(client, str) else client,
            "prime_raw": (prime or "").strip() if isinstance(prime, str) else None,
            "sub_raw": (sub or "").strip() if isinstance(sub, str) else None,
            "details": (details or "").strip() if isinstance(details, str) else None,
            "pool": (pool or "").strip() if isinstance(pool, str) else None,
            "date_submitted": parse_date_str(sub_date),
            "client_contract_number": (cc_no or "").strip() if isinstance(cc_no, str) else None,
            "msmm_contract_number": (msmm_no or "").strip() if isinstance(msmm_no, str) else None,
            "contract_expiry_date": parse_date_str(exp),
            "msmm_used": parse_money(used),
            "msmm_remaining": parse_money(remain),
            "org_type": current_org,
        })
    return out


def parse_awarded(path: Path):
    rows = _read_csv(path)
    # Find a header row: first cell 'Proposal Year'
    out: list[dict] = []
    current_org = None
    header_seen = False
    for r in rows:
        if not r:
            continue
        cell0 = (r[0] or "").strip()
        if cell0 == "Proposal Year":
            header_seen = True
            continue
        if not header_seen:
            continue
        if cell0.startswith("Org Type :"):
            current_org = cell0.split(":", 1)[1].strip()
            continue
        if not cell0:
            continue
        # data row — 16 columns like awaiting
        def cell(i):
            return r[i] if i < len(r) else ""

        out.append({
            "proposal_year": int(cell(0)) if (cell(0) or "").strip().isdigit() else None,
            "title": (cell(1) or "").strip() or None,
            "client_raw": (cell(2) or "").strip() or None,
            "prime_raw": (cell(3) or "").strip() or None,
            "sub_raw": (cell(4) or "").strip() or None,
            "stage": (cell(6) or "").strip() or None,
            "details": (cell(7) or "").strip() or None,
            "pool": (cell(8) or "").strip() or None,
            "date_submitted": parse_date_str(cell(9)),
            "client_contract_number": (cell(10) or "").strip() or None,
            "msmm_contract_number": (cell(11) or "").strip() or None,
            "contract_expiry_date": parse_date_str(cell(12)),
            "msmm_used": parse_money(cell(14)),
            "msmm_remaining": parse_money(cell(15)),
            "org_type": current_org,
        })
    return out


def parse_invoice(path: Path, year: int):
    rows = _read_csv(path)
    # Header is the row starting with 'Project No.'
    hdr_idx = next(i for i, r in enumerate(rows) if r and (r[0] or "").strip() == "Project No.")
    # Column indices (based on the 2026 file layout; 2025 has the same structure):
    #   0 Project No.
    #   1 Project Name
    #   2 PM
    #   3 Contract Amount
    #  12 True MSMM Contract Amount
    #  13 MSMM Remaining to Bill As of 1/1/YYYY
    #  14..25 monthly (Jan..Dec)  — but 2026 has May Billing Notes wedged AFTER May (idx 19)
    # 2025 file has no 'May Billing Notes' column — months are 14..25 contiguous
    # 2026 file: months 14..18 = Jan..May, 19 = May Billing Notes, 20..26 = Jun..Dec
    is_2026_layout = year == 2026

    out: list[dict] = []
    for r in rows[hdr_idx + 1:]:
        if not r:
            continue

        def cell(i):
            return r[i] if i < len(r) else ""

        project_number = (cell(0) or "").strip() or None
        project_name = (cell(1) or "").strip() or None

        # Skip blank rows
        if not project_name:
            continue
        upper_name = project_name.upper()
        # Skip section markers and totals
        if upper_name in ("ENG PROJECTS", "PM PROJECTS"):
            continue
        if upper_name.startswith("TOTAL"):
            continue
        if (project_number or "").upper().startswith("TOTAL"):
            continue
        # Skip footnotes / "* Not Included" lines
        if project_name.startswith("*") or "Not Included" in project_name:
            continue
        # Skip placeholder project codes
        if project_number and re.fullmatch(r"\d{4}[Xx]+", project_number):
            project_number = None

        contract_amount = parse_money(cell(3))
        msmm_remaining_year_start = parse_money(cell(13))

        if is_2026_layout:
            months = [parse_money(cell(14)),
                      parse_money(cell(15)),
                      parse_money(cell(16)),
                      parse_money(cell(17)),
                      parse_money(cell(18)),
                      parse_money(cell(20)),  # skip May Notes at 19
                      parse_money(cell(21)),
                      parse_money(cell(22)),
                      parse_money(cell(23)),
                      parse_money(cell(24)),
                      parse_money(cell(25)),
                      parse_money(cell(26))]
        else:
            months = [parse_money(cell(14 + i)) for i in range(12)]

        out.append({
            "year": year,
            "project_number": project_number,
            "project_name": project_name,
            "contract_amount": contract_amount,
            "type": "ENG",  # both files are the ENG export
            "msmm_remaining_to_bill_year_start": msmm_remaining_year_start,
            "jan_amount": months[0], "feb_amount": months[1], "mar_amount": months[2],
            "apr_amount": months[3], "may_amount": months[4], "jun_amount": months[5],
            "jul_amount": months[6], "aug_amount": months[7], "sep_amount": months[8],
            "oct_amount": months[9], "nov_amount": months[10], "dec_amount": months[11],
        })
    return out


# --------------------------------------------------------------------------
# Upload helpers
# --------------------------------------------------------------------------
def fetch_all(table: str, select: str = "*") -> list[dict]:
    params = {"select": select, "limit": "10000"}
    r = _req("GET", table, params=params)
    return r.json()


def insert_rows(table: str, rows: list[dict]) -> list[dict]:
    if not rows:
        return []
    r = _req("POST", table, json=rows, prefer="return=representation")
    return r.json()


def upsert_rows(table: str, rows: list[dict], on_conflict: str) -> list[dict]:
    if not rows:
        return []
    r = _req(
        "POST",
        table,
        json=rows,
        params={"on_conflict": on_conflict},
        prefer="return=representation,resolution=ignore-duplicates",
    )
    return r.json()


def delete_all(table: str, filter_col: str = "id"):
    # PostgREST requires at least one filter. Use something always-true-ish.
    if filter_col == "id":
        params = {"id": "neq.00000000-0000-0000-0000-000000000000"}
    else:
        params = {filter_col: "not.is.null"}
    r = requests.delete(f"{REST}/{table}", headers=H_BASE, params=params, timeout=60)
    if r.status_code >= 400 and r.status_code != 404:
        raise RuntimeError(f"DELETE {table} → {r.status_code}: {r.text[:300]}")


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="parse only; no writes")
    ap.add_argument("--wipe", action="store_true", help="DELETE all data first (keeps users, awarded_stages, MSMM company)")
    args = ap.parse_args()

    if not args.dry_run:
        preflight()

    # ----- Parse -----
    print("\nParsing files…")
    pot2025 = parse_potential(DATA / "Potential Projects Data/2025 Year-to-Date Contract Status.csv", 2025)
    pot2026 = parse_potential(DATA / "Potential Projects Data/2026 Year-to-Date Contract Status.csv", 2026)
    awaiting = parse_awaiting(DATA / "Awating Verdict/Year 2026 Awaiting.xlsx")
    awarded = parse_awarded(DATA / "Awarded Data/Year 2026 Awarded.csv")
    inv2025 = parse_invoice(DATA / "Invoice Cycle Data/Year 2025 Anticipated Invoice Cycle-ENG.csv", 2025)
    inv2026 = parse_invoice(DATA / "Invoice Cycle Data/Year 2026Anticipated Invoice Cycle-ENG.csv", 2026)

    print(f"  Potential 2025: {len(pot2025)} rows")
    print(f"  Potential 2026: {len(pot2026)} rows")
    print(f"  Awaiting Verdict: {len(awaiting)} rows")
    print(f"  Awarded: {len(awarded)} rows")
    print(f"  Invoice 2025: {len(inv2025)} rows")
    print(f"  Invoice 2026: {len(inv2026)} rows")

    # ----- Apply 2026 potential fixes in-memory -----
    for row in pot2026:
        name = row["project_name"]
        if name in POTENTIAL_2026_SUB_FIXES:
            client_name, client_district, org_type, prime = POTENTIAL_2026_SUB_FIXES[name]
            row["client_canonical"] = (client_name, client_district)
            row["client_org_type"] = org_type
            row["prime_company_name"] = prime
            row["role"] = "Sub"
        elif (not row["role_raw"]) and (not row["client_raw"]):
            # Blank role + blank client → try inference
            inferred = POTENTIAL_BLANK_CLIENT_INFERENCES.get(name)
            if inferred:
                client_name, client_district, org_type = inferred
                row["client_canonical"] = (client_name, client_district)
                row["client_org_type"] = org_type
            else:
                row["client_canonical"] = (None, None)
                row["client_org_type"] = None
            row["prime_company_name"] = None
            row["role"] = None
        else:
            # Normal case: role explicit, client column trustworthy
            n, d = canonicalize_client(row["client_raw"])
            row["client_canonical"] = (n, d)
            row["client_org_type"] = None  # filled later from awarded/awaiting cross-reference
            row["prime_company_name"] = None
            row["role"] = row["role_raw"] or None
            # If role='Sub' but we didn't apply a fix above, the client column is still the prime.
            # That's a bug in source data not covered by known fixes; warn but proceed.
            if row["role"] == "Sub":
                print(f"  WARN 2026 Potential Sub row without fix: {name!r} — client_raw may be a prime firm.")

    for row in pot2025:
        # User: MSMM prime for all 2025; no client column; subs are free-text.
        row["client_canonical"] = (None, None)
        row["client_org_type"] = None
        row["prime_company_name"] = None
        row["role"] = "Prime"

    # ----- Build client + company dedup sets -----
    # Org type mapping: prefer Awarded section then Awaiting section then Potential inference
    client_org_type: dict[tuple, str] = {}

    def remember_client(raw, org_type):
        n, d = canonicalize_client(raw)
        if not n:
            return None
        key = (n, d)
        if org_type and key not in client_org_type:
            client_org_type[key] = org_type
        return key

    clients_set: set[tuple] = set()
    companies_set: set[str] = set()

    for row in awarded:
        k = remember_client(row["client_raw"], row["org_type"])
        if k: clients_set.add(k)
        if row["prime_raw"]:
            companies_set.add(row["prime_raw"].strip())
        for s in split_semi_list(row["sub_raw"]):
            companies_set.add(s)

    for row in awaiting:
        k = remember_client(row["client_raw"], row["org_type"])
        if k: clients_set.add(k)
        if row["prime_raw"]:
            companies_set.add(row["prime_raw"].strip())
        for s in split_semi_list(row["sub_raw"]):
            companies_set.add(s)

    for row in pot2026 + pot2025:
        k = row.get("client_canonical") or (None, None)
        if k[0]:
            clients_set.add(k)
            if row.get("client_org_type") and k not in client_org_type:
                client_org_type[k] = row["client_org_type"]
        if row.get("prime_company_name"):
            companies_set.add(row["prime_company_name"])

    companies_set.discard("")
    companies_set.discard(None)  # type: ignore

    print(f"\nUnique clients: {len(clients_set)}")
    print(f"Unique companies: {len(companies_set)}")

    if args.dry_run:
        print("\n--dry-run: no writes. Sample parsed rows:")
        print("  potential 2026[0]:", pot2026[0] if pot2026 else None)
        print("  awaiting[0]:", awaiting[0] if awaiting else None)
        print("  awarded[0]:", awarded[0] if awarded else None)
        print("  invoice 2026[0]:", inv2026[0] if inv2026 else None)
        print("  clients sample:", sorted(clients_set)[:5])
        print("  companies sample:", sorted(companies_set)[:10])
        print("  org types:", {k: v for k, v in list(client_org_type.items())[:5]})
        return

    # ----- Wipe (optional) -----
    if args.wipe:
        print("\nWiping existing data (children first)…")
        for t in [
            "anticipated_invoice_pms", "anticipated_invoice",
            "closed_out_project_pms", "closed_out_projects",
            "awarded_project_subs", "awarded_project_pms", "awarded_projects",
            "awaiting_verdict_subs", "awaiting_verdict_pms", "awaiting_verdict",
            "potential_project_subs", "potential_project_pms", "potential_projects",
            "event_attendees", "events",
            "alert_recipients", "alert_fires", "alerts",
        ]:
            pk = "id"
            if t.endswith("_pms") or t.endswith("_subs") or t in ("event_attendees", "alert_recipients"):
                # Composite-PK tables — use any non-null column
                if t.startswith("potential_project_subs"):
                    pk = "potential_project_id"
                elif t.startswith("potential_project_pms"):
                    pk = "potential_project_id"
                elif t.startswith("awaiting_verdict_subs") or t.startswith("awaiting_verdict_pms"):
                    pk = "awaiting_verdict_id"
                elif t.startswith("awarded_project_subs") or t.startswith("awarded_project_pms"):
                    pk = "awarded_project_id"
                elif t.startswith("closed_out_project_pms"):
                    pk = "closed_out_project_id"
                elif t.startswith("anticipated_invoice_pms"):
                    pk = "anticipated_invoice_id"
                elif t == "event_attendees":
                    pk = "event_id"
                elif t == "alert_recipients":
                    pk = "alert_id"
            delete_all(t, pk)
        # clients (no FK dependents after the above) and non-MSMM companies
        delete_all("clients", "id")
        r = requests.delete(
            f"{REST}/companies",
            headers=H_BASE,
            params={"is_msmm": "eq.false"},
            timeout=60,
        )
        if r.status_code >= 400 and r.status_code != 404:
            raise RuntimeError(f"DELETE companies: {r.status_code}: {r.text[:300]}")
        print("  wiped.")

    # ----- Upsert companies -----
    print(f"\nUpserting {len(companies_set)} companies…")
    companies_payload = [{"name": n} for n in sorted(companies_set)]
    upsert_rows("companies", companies_payload, on_conflict="name")
    companies_all = fetch_all("companies", select="id,name")
    comp_id = {c["name"]: c["id"] for c in companies_all}
    msmm_id = comp_id.get("MSMM")
    if not msmm_id:
        sys.exit("MSMM company row missing — was the migration applied?")

    # ----- Upsert clients -----
    print(f"Upserting {len(clients_set)} clients…")
    clients_payload = []
    for (n, d) in sorted(clients_set, key=lambda x: (x[0], x[1] or "")):
        clients_payload.append({
            "name": n,
            "district": d,
            "org_type": client_org_type.get((n, d)),
        })
    # clients has an expression-based unique index (coalesce(district,'')) which
    # PostgREST's on_conflict can't match. Insert per row and tolerate 409/23505.
    inserted = skipped = 0
    for c in clients_payload:
        rr = requests.post(
            f"{REST}/clients",
            headers={**H_JSON, "Prefer": "return=representation"},
            json=c, timeout=30,
        )
        if rr.status_code < 300:
            inserted += 1
            continue
        if rr.status_code == 409 or "duplicate key" in rr.text or "23505" in rr.text:
            skipped += 1
            continue
        print(f"    WARN client insert failed for {c}: {rr.status_code} {rr.text[:200]}")
    print(f"  inserted={inserted} skipped(dup)={skipped}")
    clients_all = fetch_all("clients", select="id,name,district")
    client_id = {(c["name"], c["district"]): c["id"] for c in clients_all}

    # ----- Insert potential_projects -----
    print(f"\nInserting potential projects (2025+2026)…")
    stages_all = fetch_all("awarded_stages", select="id,name")
    stage_id_by_name = {s["name"]: s["id"] for s in stages_all}

    def client_fk(row):
        c = row.get("client_canonical") or (None, None)
        if not c[0]:
            return None
        return client_id.get(c)

    def prime_fk(row):
        name = row.get("prime_company_name")
        return comp_id.get(name) if name else None

    pot_payload = []
    for row in pot2026 + pot2025:
        pot_payload.append({
            "year": row["year"],
            "project_name": row["project_name"],
            "role": row["role"],
            "client_id": client_fk(row),
            "prime_company_id": prime_fk(row),
            "total_contract_amount": row["total_contract_amount"],
            "msmm_amount": row["msmm_amount"],
            "notes": row["notes"],
            "next_action_date": row["next_action_date"],
            "next_action_note": row["next_action_note"],
            "project_number": row["project_number"],
            "probability": row["probability"],
        })
    inserted_pot = insert_rows("potential_projects", pot_payload)
    print(f"  inserted {len(inserted_pot)} potential_projects")
    pot_id_by_key = {(p["year"], p["project_name"]): p["id"] for p in inserted_pot}

    # Subs for potential (from subs_raw — company_id stays NULL per agreement)
    subs_payload = []
    for row in pot2026 + pot2025:
        pid = pot_id_by_key.get((row["year"], row["project_name"]))
        if not pid:
            continue
        for idx, raw in enumerate(row["subs_raw"], start=1):
            amount, disc = parse_sub_cell(raw)
            if amount is None and disc is None:
                continue
            subs_payload.append({
                "potential_project_id": pid,
                "ord": idx,
                "company_id": None,
                "discipline": disc,
                "amount": amount,
            })
    if subs_payload:
        inserted_subs = insert_rows("potential_project_subs", subs_payload)
        print(f"  inserted {len(inserted_subs)} potential_project_subs")

    # ----- Insert awaiting_verdict -----
    print(f"\nInserting awaiting verdict rows…")
    aw_payload = []
    for row in awaiting:
        cname, cdist = canonicalize_client(row["client_raw"])
        aw_payload.append({
            "year": row["proposal_year"],
            "project_name": row["title"],
            "client_id": client_id.get((cname, cdist)) if cname else None,
            "prime_company_id": comp_id.get(row["prime_raw"]) if row["prime_raw"] else None,
            "project_number": None,
            "notes": None,
            "date_submitted": row["date_submitted"],
            "client_contract_number": row["client_contract_number"],
            "msmm_contract_number": row["msmm_contract_number"],
            "msmm_used": row["msmm_used"],
            "msmm_remaining": row["msmm_remaining"],
        })
    inserted_aw = insert_rows("awaiting_verdict", aw_payload)
    print(f"  inserted {len(inserted_aw)} awaiting_verdict")
    aw_id_by_title = {a["project_name"]: a["id"] for a in inserted_aw}

    aw_subs_payload = []
    for row in awaiting:
        aid = aw_id_by_title.get(row["title"])
        if not aid:
            continue
        for s in split_semi_list(row["sub_raw"]):
            cid = comp_id.get(s)
            if cid:
                aw_subs_payload.append({"awaiting_verdict_id": aid, "company_id": cid})
    # Dedup composite PK within payload
    aw_subs_payload = list({(p["awaiting_verdict_id"], p["company_id"]): p for p in aw_subs_payload}.values())
    if aw_subs_payload:
        insert_rows("awaiting_verdict_subs", aw_subs_payload)
        print(f"  inserted {len(aw_subs_payload)} awaiting_verdict_subs")

    # ----- Insert awarded_projects -----
    print(f"\nInserting awarded projects…")
    aw_proj_payload = []
    for row in awarded:
        cname, cdist = canonicalize_client(row["client_raw"])
        aw_proj_payload.append({
            "year": row["proposal_year"],
            "project_name": row["title"],
            "client_id": client_id.get((cname, cdist)) if cname else None,
            "prime_company_id": comp_id.get(row["prime_raw"]) if row["prime_raw"] else None,
            "project_number": None,
            "date_submitted": row["date_submitted"],
            "client_contract_number": row["client_contract_number"],
            "msmm_contract_number": row["msmm_contract_number"],
            "msmm_used": row["msmm_used"],
            "msmm_remaining": row["msmm_remaining"],
            "stage_id": stage_id_by_name.get(row["stage"]),
            "details": row["details"],
            "pool": row["pool"],
            "contract_expiry_date": row["contract_expiry_date"],
        })
    inserted_awp = insert_rows("awarded_projects", aw_proj_payload)
    print(f"  inserted {len(inserted_awp)} awarded_projects")
    # Key by (year, title) to handle duplicate titles across rows (same proposal, different primes)
    awp_rows_by_key = {}
    for p in inserted_awp:
        awp_rows_by_key.setdefault((p["year"], p["project_name"]), []).append(p["id"])

    awp_subs_payload = []
    # Walk awarded in the same order so we can zip source rows to inserted rows
    for src, ins in zip(awarded, inserted_awp):
        aid = ins["id"]
        for s in split_semi_list(src["sub_raw"]):
            cid = comp_id.get(s)
            if cid:
                awp_subs_payload.append({"awarded_project_id": aid, "company_id": cid})
    awp_subs_payload = list({(p["awarded_project_id"], p["company_id"]): p for p in awp_subs_payload}.values())
    if awp_subs_payload:
        insert_rows("awarded_project_subs", awp_subs_payload)
        print(f"  inserted {len(awp_subs_payload)} awarded_project_subs")

    # ----- Insert anticipated_invoice -----
    print(f"\nInserting invoice rows…")
    inv_payload = inv2025 + inv2026
    inserted_inv = insert_rows("anticipated_invoice", inv_payload)
    print(f"  inserted {len(inserted_inv)} anticipated_invoice")

    print("\nDone.")


if __name__ == "__main__":
    main()
