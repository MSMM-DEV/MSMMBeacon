#!/usr/bin/env python3
"""
MSMM Beacon — BD.xlsx → beacon.events ingester.

Parses Data/Events/BD.xlsx (single sheet, five category columns) and
upserts rows into beacon.events. Type + status are driven by column
position + cell formatting:

  Columns (1-indexed in the sheet)
    B        Projects     → type='Project',  no date, no status
    D / E    Events       → type='Event',    date+title
    G / H    Partners     → type='Partner',  date+title
    J / K    AI           → type='AI',       date+title
    M / N    Meetings     → type='Meetings', date+title

  Formatting on the date cell (falls back to title cell if date blank)
    red font (FFFF0000) → status='Happened'
    yellow fill (FFFFFF00) → status='Booked'
    neither              → status=NULL  (future/tentative)
    literal '?' in date cell → event_date=NULL (date unknown)

Dedupe against existing beacon.events:
  • same type
  • same event_date (both NULL counts as equal)
  • normalized title match (lowercase, collapsed ws, trailing " 1 2" /
    digit-suffix stripped)

Rows where (type, date) collide with an existing DB row but the
normalized titles differ are NOT auto-inserted — they print as
DATE-COLLISION so the operator can decide (usually the DB has a
cleaner-spelled title like 'Banquet' vs the xlsx 'Banq').

Usage
    python3 scripts/ingest_bd_events.py              # dry-run
    python3 scripts/ingest_bd_events.py --commit     # actually insert
    python3 scripts/ingest_bd_events.py --commit --force
        # also insert the DATE-COLLISION rows (not recommended)

Requires SUPABASE_URL and SUPABASE_SERVICE_KEY in .env.
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
load_dotenv(ROOT / ".env")

URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
KEY = os.environ.get("SUPABASE_SERVICE_KEY", "")
if not URL or not KEY:
    sys.exit("Missing SUPABASE_URL / SUPABASE_SERVICE_KEY in .env")

REST = f"{URL}/rest/v1"
H = {
    "apikey": KEY,
    "Authorization": f"Bearer {KEY}",
    "Accept-Profile": "beacon",
    "Content-Profile": "beacon",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}

XLSX = ROOT / "Data" / "Events" / "BD.xlsx"

# Column layout inside BD.xlsx (1-indexed). Dates + titles share a column
# pair; Projects is title-only.
COLUMNS = [
    ("Project",  None, 2),   # title in col B, no date
    ("Event",    4,    5),   # date D, title E
    ("Partner",  7,    8),   # date G, title H
    ("AI",       10,   11),  # date J, title K
    ("Meetings", 13,   14),  # date M, title N
]
DATA_ROWS = range(3, 13)     # rows 3..12; row 2 is headers, 14-15 are legend

RED_FONT   = "FFFF0000"
YELLOW_FILL = "FFFFFF00"


# ---------------------------------------------------------------------------
# xlsx parsing
# ---------------------------------------------------------------------------

def _rgb(colorish) -> str | None:
    """Return the hex RGB string on a Font.color / Fill.fgColor, or None."""
    if not colorish:
        return None
    try:
        rgb = colorish.rgb
    except AttributeError:
        return None
    if isinstance(rgb, str) and rgb and rgb != "00000000":
        return rgb
    return None


def _classify(cell) -> str | None:
    """Return 'Happened' (red font), 'Booked' (yellow fill), or None."""
    if cell is None:
        return None
    if cell.font and _rgb(cell.font.color) == RED_FONT:
        return "Happened"
    fill = cell.fill
    if fill and _rgb(fill.fgColor) == YELLOW_FILL:
        return "Booked"
    return None


def _as_iso(v) -> str | None:
    """Coerce a cell value to ISO date string, or None for blanks / '?'."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if s in ("", "?"):
        return None
    # Fall back to a loose parse (shouldn't happen with openpyxl dates).
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def parse_bd() -> list[dict]:
    wb = load_workbook(XLSX, data_only=True)
    ws = wb.active

    rows: list[dict] = []
    for r in DATA_ROWS:
        for ev_type, date_col, title_col in COLUMNS:
            title_cell = ws.cell(row=r, column=title_col)
            date_cell  = ws.cell(row=r, column=date_col) if date_col else None

            title = title_cell.value
            if title is None or str(title).strip() == "":
                continue

            iso = _as_iso(date_cell.value) if date_cell is not None else None

            # Prefer status color on the date cell (the dated cell), fall
            # back to the title cell so Projects + '?' rows still classify.
            status = _classify(date_cell) or _classify(title_cell)

            rows.append({
                "type":   ev_type,
                "title":  str(title).strip(),
                "date":   iso,
                "status": status,
                "_src":   f"row{r} col{title_col}",
            })
    return rows


# ---------------------------------------------------------------------------
# dedupe
# ---------------------------------------------------------------------------

def _norm(title: str) -> str:
    t = title.lower().strip()
    t = re.sub(r"\s+", " ", t)
    t = re.sub(r"[^a-z0-9 ]", "", t)
    # Trailing numeric markers like " 1 2" are pagination refs in the
    # spreadsheet ("Coastal Day 1 2"), not part of the canonical title.
    t = re.sub(r"(?:\s+\d+)+$", "", t).strip()
    return t


def fetch_existing() -> list[dict]:
    r = requests.get(
        f"{REST}/events",
        headers=H,
        params={"select": "id,event_date,type,title,status"},
        timeout=15,
    )
    r.raise_for_status()
    return r.json()


def plan(bd_rows: list[dict], db_rows: list[dict]) -> dict[str, list]:
    by_type_date: dict[tuple[str, str | None], list[dict]] = {}
    by_exact: set[tuple[str, str | None, str]] = set()
    for d in db_rows:
        key = (d["type"], d.get("event_date"))
        by_type_date.setdefault(key, []).append(d)
        by_exact.add((d["type"], d.get("event_date"), _norm(d["title"])))

    inserts, skip_dup, collisions = [], [], []
    for row in bd_rows:
        exact_key = (row["type"], row["date"], _norm(row["title"]))
        if exact_key in by_exact:
            skip_dup.append(row)
            continue
        matches = by_type_date.get((row["type"], row["date"]), [])
        dated_collision = row["date"] is not None and matches
        if dated_collision:
            collisions.append({"bd": row, "db_matches": matches})
            continue
        inserts.append(row)
    return {"insert": inserts, "skip_dup": skip_dup, "collision": collisions}


# ---------------------------------------------------------------------------
# commit
# ---------------------------------------------------------------------------

def to_payload(row: dict) -> dict:
    return {
        "type":       row["type"],
        "title":      row["title"],
        "event_date": row["date"],
        "status":     row["status"],
    }


def insert_batch(rows: list[dict]) -> list[dict]:
    if not rows:
        return []
    r = requests.post(
        f"{REST}/events",
        headers=H,
        json=[to_payload(row) for row in rows],
        timeout=30,
    )
    if not r.ok:
        sys.exit(f"Insert failed: {r.status_code} {r.text}")
    return r.json()


# ---------------------------------------------------------------------------
# reporting
# ---------------------------------------------------------------------------

def _fmt(row: dict) -> str:
    d = row.get("date") or row.get("event_date") or "    —    "
    s = row.get("status") or "—"
    return f"{row['type']:<9} {d:<12} [{s:<8}] {row['title']}"


def print_plan(p: dict) -> None:
    print(f"Parsed {sum(len(v) for v in p.values())} candidate rows from {XLSX.name}")
    print()

    print(f"INSERT ({len(p['insert'])})")
    for row in p["insert"]:
        print("  + " + _fmt(row))
    print()

    print(f"SKIP — exact duplicate of existing beacon.events row ({len(p['skip_dup'])})")
    for row in p["skip_dup"]:
        print("  = " + _fmt(row))
    print()

    print(f"DATE-COLLISION — (type,date) exists in DB with a different title ({len(p['collision'])})")
    for item in p["collision"]:
        print("  ! BD : " + _fmt(item["bd"]))
        for m in item["db_matches"]:
            print("    DB : " + _fmt(m))
    print()


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--commit", action="store_true",
                    help="actually insert (default is dry-run)")
    ap.add_argument("--force", action="store_true",
                    help="also insert DATE-COLLISION rows (not recommended)")
    args = ap.parse_args()

    bd_rows = parse_bd()
    db_rows = fetch_existing()
    p = plan(bd_rows, db_rows)
    print_plan(p)

    if not args.commit:
        print("dry-run only; re-run with --commit to insert.")
        return

    to_insert = list(p["insert"])
    if args.force:
        to_insert.extend(item["bd"] for item in p["collision"])

    inserted = insert_batch(to_insert)
    print(f"inserted {len(inserted)} row(s) into beacon.events.")


if __name__ == "__main__":
    main()
