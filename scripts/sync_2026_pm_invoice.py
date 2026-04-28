#!/usr/bin/env python3
"""
Wipe all (year=2026, type='PM') anticipated_invoice rows and reinsert
them from the `2026` worksheet of the PM-side workbook so the table
matches the spreadsheet exactly.

Default source (override with --xlsx):
  ~/Library/CloudStorage/Egnyte-msmm/Shared/PData/0_StatusSS/
    0MSMM_2026 Anticipated Invoice Cycle-PM.xlsx

Layout notes (observed in that workbook, sheet "2026"):
  • Row 4  = header (Project No., Project Name, PM, Contract Amount, …).
  • Row 5  = "PM Projects" section marker.
  • Row 6+ = data rows; rows whose Project Name starts with "TOTAL" or
    is blank are skipped.

Column map (1-indexed, contiguous — no skip column):
  B=2   Project No.
  C=3   Project Name
  D=4   PM (single name; resolved via the same PM_OVERRIDES + roster
        match as scripts/sync_2026_invoice.py)
  E=5   Contract Amount
  F=6   MSMM Remaining to Bill in Future
  G=7..R=18   Jan..Dec MSMM Billing
  S=19  2026 MSMM Total Actual   (persisted as ytd_actual_override)
  T=20  2027 MSMM Rollforward    (persisted as rollforward_override)

Why scope the wipe to type='PM'?
  scripts/sync_2026_invoice.py wipes year=2026 unconditionally to land
  the ENG sheet. If we did the same here, running either script would
  trample the other type's rows. Scoping by type keeps the two scripts
  independent — run them in either order.

This is destructive for (year=2026, type='PM') invoice rows (and their
*_pms join rows). Use --dry-run to preview without writing.
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()
SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_SERVICE_KEY"]

DEFAULT_XLSX = Path(
    "/Users/rajmehta/Library/CloudStorage/Egnyte-msmm/Shared/PData/"
    "0_StatusSS/0MSMM_2026 Anticipated Invoice Cycle-PM.xlsx"
)

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Accept-Profile": "beacon",
    "Content-Profile": "beacon",
    "Prefer": "return=representation",
}

MONTH_COLS = [
    "jan_amount", "feb_amount", "mar_amount", "apr_amount",
    "may_amount", "jun_amount", "jul_amount", "aug_amount",
    "sep_amount", "oct_amount", "nov_amount", "dec_amount",
]

# Same convention as sync_2026_invoice.py — "Scott" alone is ambiguous
# between Scott Chehardy and Scott Douglas; the sheet uses "Scott D."
# for Douglas (which the roster's short_name already matches), so the
# override only needs the bare-name fallback.
PM_OVERRIDES = {
    "Scott": "Scott Douglas",
}


def api(method: str, path: str, **kw):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    r = requests.request(method, url, headers=HEADERS, **kw)
    if not r.ok:
        raise RuntimeError(f"{method} {path} → {r.status_code}: {r.text}")
    return r.json() if r.text else []


def num(v):
    if v is None or v == "":
        return 0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0


def override_num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_worksheet(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["2026"]
    out: list[dict] = []
    for i in range(6, ws.max_row + 1):
        proj_no = ws.cell(row=i, column=2).value
        name = ws.cell(row=i, column=3).value
        pm = ws.cell(row=i, column=4).value
        contract = ws.cell(row=i, column=5).value
        msmm_rem = ws.cell(row=i, column=6).value
        months = [ws.cell(row=i, column=c).value for c in range(7, 19)]
        ytd_override = override_num(ws.cell(row=i, column=19).value)
        rf_override = override_num(ws.cell(row=i, column=20).value)

        if not name or not isinstance(name, str):
            continue
        up = name.upper()
        if up in ("ENG PROJECTS", "PM PROJECTS"):
            continue
        if up.startswith("TOTAL"):
            continue
        if name.startswith("*") or "Not Included" in name:
            continue

        if isinstance(proj_no, str) and re.fullmatch(r"\d{4}[Xx]+", proj_no):
            proj_no = None
        elif proj_no is not None:
            proj_no = str(proj_no).strip() or None

        out.append({
            "row": i,
            "project_number": proj_no,
            "project_name": name.strip(),
            "pm": (pm or "").strip() if isinstance(pm, str) else (str(pm) if pm else ""),
            "contract_amount": num(contract),
            "msmm_remaining_to_bill_year_start": num(msmm_rem),
            "months": [num(m) for m in months],
            "ytd_actual_override": ytd_override,
            "rollforward_override": rf_override,
        })
    return out


def resolve_pm(name: str, users: list[dict]) -> str | None:
    if not name:
        return None
    probe = PM_OVERRIDES.get(name, name)
    parts = probe.split()
    for u in users:
        short = (u.get("short_name") or "").lower()
        first = (u.get("first_name") or "").lower()
        last = (u.get("last_name") or "").lower()
        disp = (u.get("display_name") or "").lower()
        if short and short == probe.lower():
            return u["id"]
        if disp == probe.lower():
            return u["id"]
        if len(parts) >= 2 and first == parts[0].lower() and last.startswith(parts[-1].lower()):
            return u["id"]
        if len(parts) == 1 and first == parts[0].lower():
            candidates = [x for x in users if (x.get("first_name") or "").lower() == parts[0].lower()]
            if len(candidates) == 1:
                return u["id"]
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="Parse + plan without writing.")
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="Override workbook path.")
    args = ap.parse_args()

    if not args.xlsx.exists():
        print(f"! Workbook not found: {args.xlsx}", file=sys.stderr)
        sys.exit(2)

    print(f"Loading {args.xlsx.name}…")
    rows = parse_worksheet(args.xlsx)
    print(f"  parsed {len(rows)} PM rows")
    for r in rows:
        months_str = ", ".join(f"{m:,.0f}" if m else "·" for m in r["months"])
        print(f"  · #{r['project_number'] or '——'} {r['project_name']}")
        print(f"      PM={r['pm']!r}  contract=${r['contract_amount']:,.2f}  remaining=${r['msmm_remaining_to_bill_year_start']:,.2f}")
        print(f"      months: [{months_str}]")
        print(f"      ytd_override={r['ytd_actual_override']}  rf_override={r['rollforward_override']}")

    if not rows:
        print("No PM rows found — nothing to do.")
        return

    print("\nFetching user roster for PM resolution…")
    users = api("GET", "users?select=id,first_name,last_name,display_name,short_name")
    print(f"  {len(users)} users")

    pm_map: dict[str, str | None] = {}
    for r in rows:
        if r["pm"] and r["pm"] not in pm_map:
            pm_map[r["pm"]] = resolve_pm(r["pm"], users)
    missing = [k for k, v in pm_map.items() if v is None and k]
    if missing:
        print(f"  ! PM names not matched to roster: {missing}")
    else:
        for k, v in pm_map.items():
            u = next((x for x in users if x["id"] == v), None)
            label = u.get("display_name") if u else "?"
            print(f"  · '{k}' → {label}")

    print("\nExisting (year=2026, type='PM') invoices in DB:")
    existing = api("GET", "anticipated_invoice?year=eq.2026&type=eq.PM&select=id,project_name,project_number")
    for e in existing:
        print(f"  · #{e.get('project_number') or '——'} {e.get('project_name')}  (id={e['id']})")
    print(f"  {len(existing)} rows (will be replaced)")

    if args.dry_run:
        print("\n--dry-run: no writes performed.")
        return

    if existing:
        print("\nDeleting existing PM 2026 invoice rows…")
        api("DELETE", "anticipated_invoice?year=eq.2026&type=eq.PM")

    print("Inserting fresh PM invoice rows…")
    payloads = []
    for r in rows:
        p = {
            "year": 2026,
            "project_name": r["project_name"],
            "project_number": r["project_number"],
            "contract_amount": r["contract_amount"],
            "type": "PM",
            "msmm_remaining_to_bill_year_start": r["msmm_remaining_to_bill_year_start"],
            "source_potential_id": None,
            "ytd_actual_override": r["ytd_actual_override"],
            "rollforward_override": r["rollforward_override"],
        }
        for col, v in zip(MONTH_COLS, r["months"]):
            p[col] = v
        payloads.append(p)

    inserted = api("POST", "anticipated_invoice", json=payloads)
    print(f"  inserted {len(inserted)}")

    pm_rows = []
    for inv_row, src in zip(inserted, rows):
        uid = pm_map.get(src["pm"])
        if uid:
            pm_rows.append({"anticipated_invoice_id": inv_row["id"], "user_id": uid})
    if pm_rows:
        api("POST", "anticipated_invoice_pms", json=pm_rows)
        print(f"  inserted {len(pm_rows)} PM tags")

    print("\n✓ PM sync complete.")


if __name__ == "__main__":
    main()
