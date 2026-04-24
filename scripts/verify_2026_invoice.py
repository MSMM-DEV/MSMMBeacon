#!/usr/bin/env python3
"""
Cell-by-cell diff between the xlsx "2026" worksheet and the Supabase
anticipated_invoice table. Prints every discrepancy it finds. No writes.

Use this to:
  • confirm that a sync actually landed correctly,
  • catch columns the sync script forgot to copy,
  • inspect PM resolution.
"""
from __future__ import annotations

import os
import re
import sys
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()
URL = os.environ["SUPABASE_URL"]
KEY = os.environ["SUPABASE_SERVICE_KEY"]
REPO = Path(__file__).resolve().parent.parent
XLSX = REPO / "Data" / "Invoice Cycle Data" / "NEW_2026.xlsx"
H = {
    "apikey": KEY,
    "Authorization": f"Bearer {KEY}",
    "Accept-Profile": "beacon",
}

ORANGE_RGB = "FFFFC000"
MONTH_LABELS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_COLS = ["jan_amount", "feb_amount", "mar_amount", "apr_amount",
              "may_amount", "jun_amount", "jul_amount", "aug_amount",
              "sep_amount", "oct_amount", "nov_amount", "dec_amount"]


def num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def approx_eq(a, b):
    """Compare numbers allowing for float-precision drift."""
    a = num(a)
    b = num(b)
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return abs(a - b) < 0.005


def fill_rgb(cell):
    f = cell.fill
    if not f or not f.fgColor:
        return None
    try:
        rgb = f.fgColor.rgb
        if isinstance(rgb, str) and rgb != "00000000":
            return rgb
    except Exception:
        pass
    return None


def parse_xlsx():
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["2026"]
    rows = []
    for i in range(6, ws.max_row + 1):
        proj_no = ws.cell(row=i, column=2).value
        name    = ws.cell(row=i, column=3).value
        pm      = ws.cell(row=i, column=4).value
        if not name or not isinstance(name, str):
            continue
        up = name.upper()
        if up in ("ENG PROJECTS", "PM PROJECTS") or up.startswith("TOTAL"):
            continue
        if name.startswith("*") or "Not Included" in name:
            continue
        if isinstance(proj_no, str) and re.fullmatch(r"\d{4}[Xx]+", proj_no):
            proj_no = None
        elif proj_no is not None:
            proj_no = str(proj_no).strip() or None

        contract = ws.cell(row=i, column=5).value
        msmm_rem = ws.cell(row=i, column=14).value
        months   = [ws.cell(row=i, column=c).value for c in (15,16,17,18,19,21,22,23,24,25,26,27)]
        ytd      = ws.cell(row=i, column=28).value  # AB
        rf       = ws.cell(row=i, column=29).value  # AC
        fill     = fill_rgb(ws.cell(row=i, column=3))

        rows.append({
            "row": i,
            "project_number": proj_no,
            "project_name": name.strip(),
            "pm": (pm or "").strip() if isinstance(pm, str) else "",
            "contract_amount": num(contract),
            "msmm_remaining_to_bill_year_start": num(msmm_rem),
            "months": [num(m) for m in months],
            "ytd_actual_override": num(ytd),
            "rollforward_override": num(rf),
            "is_orange": fill == ORANGE_RGB,
        })
    return rows


def fetch_db():
    r = requests.get(
        f"{URL}/rest/v1/anticipated_invoice",
        params={
            "year": "eq.2026",
            "select": "id,project_number,project_name,contract_amount,type,"
                      "msmm_remaining_to_bill_year_start,source_potential_id,"
                      "ytd_actual_override,rollforward_override,"
                      + ",".join(MONTH_COLS)
                      + ",pms:anticipated_invoice_pms(user_id)",
        },
        headers=H,
    )
    r.raise_for_status()
    return r.json()


def match_key(row):
    """Match xlsx row ↔ db row by project_number when available,
    otherwise by project_name (trimmed)."""
    pn = (row.get("project_number") or "").strip()
    return pn if pn else (row.get("project_name") or "").strip()


def main():
    xrows = parse_xlsx()
    dbrows = fetch_db()

    print(f"xlsx: {len(xrows)} rows | db: {len(dbrows)} rows")

    # Index by match_key — prefer number, fall back to name.
    dbmap_num = {r["project_number"]: r for r in dbrows if r.get("project_number")}
    dbmap_name = {r["project_name"]: r for r in dbrows}

    missing_in_db = []
    extra_in_db = set(r["id"] for r in dbrows)
    mismatches = []

    for x in xrows:
        pn = (x.get("project_number") or "").strip()
        key = pn or x["project_name"]
        db = dbmap_num.get(pn) if pn else None
        if not db:
            db = dbmap_name.get(x["project_name"])
        if not db:
            missing_in_db.append(key)
            continue
        extra_in_db.discard(db["id"])

        def diff(col, xval, dbval, fmt=str):
            if isinstance(xval, float) or isinstance(dbval, float) or col.endswith("_amount") or col.endswith("_override") or col in MONTH_COLS:
                if not approx_eq(xval, dbval):
                    mismatches.append((key, col, fmt(xval), fmt(dbval)))
            else:
                if (xval or None) != (dbval or None):
                    mismatches.append((key, col, fmt(xval), fmt(dbval)))

        diff("project_number", x["project_number"], db.get("project_number"))
        diff("project_name",   x["project_name"],   db.get("project_name"))
        diff("contract_amount", x["contract_amount"], db.get("contract_amount"))
        diff("msmm_remaining_to_bill_year_start",
             x["msmm_remaining_to_bill_year_start"],
             db.get("msmm_remaining_to_bill_year_start"))
        for i, col in enumerate(MONTH_COLS):
            diff(col, x["months"][i], db.get(col))
        diff("ytd_actual_override",   x["ytd_actual_override"],   db.get("ytd_actual_override"))
        diff("rollforward_override",  x["rollforward_override"],  db.get("rollforward_override"))
        # Orange: expect source_potential_id non-null if and only if fill=orange
        want_orange = x["is_orange"]
        has_source  = db.get("source_potential_id") is not None
        if want_orange != has_source:
            mismatches.append((key, "orange_link",
                               "linked" if want_orange else "none",
                               "linked" if has_source else "none"))

    # Extras: DB rows not in xlsx
    extras = [r for r in dbrows if r["id"] in extra_in_db]

    print()
    print(f"Missing in DB  : {len(missing_in_db)}")
    for k in missing_in_db:
        print("   -", k)
    print(f"Extra in DB    : {len(extras)}")
    for r in extras:
        print("   +", r.get("project_number") or r.get("project_name"))
    print(f"Cell mismatches: {len(mismatches)}")
    if mismatches:
        print(f"  {'project':38} {'column':32} {'xlsx':>16}  {'db':>16}")
        for k, c, xv, dv in mismatches[:120]:
            print(f"  {str(k)[:38]:38} {c:32} {str(xv):>16}  {str(dv):>16}")

    if not missing_in_db and not extras and not mismatches:
        print("\n✓ xlsx and DB match exactly.")


if __name__ == "__main__":
    main()
