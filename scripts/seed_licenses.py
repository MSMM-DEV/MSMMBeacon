#!/usr/bin/env python3
"""
Seed beacon_v2.licenses from the licenses-module spreadsheet.

Maps the source columns (LIC_ID / LIC_NAME / LIC_STATE / LIC_TYPE / LIC_NO /
ASCEM_NO / FIRST_ISSUE_DATE / EXPIRATION_DATE / LIC_NOTIFY_NAMES / EMAIL_ENABLED
/ LIC_COMMENTS) into beacon_v2.licenses. Nulls (no expiry / no license no) are
preserved; '_x000D_' carriage-return artifacts are stripped from comments;
notify addresses split on commas into a text[].

Idempotent: upserts on legacy_id (the original LIC_ID), so re-running updates in
place rather than duplicating. Requires migration 20260609120000 applied first.

Usage:
    python3 scripts/seed_licenses.py            # dry-run (read-only)
    python3 scripts/seed_licenses.py --apply    # upsert into beacon_v2.licenses
"""
from __future__ import annotations

import datetime as dt
import os
import sys

import openpyxl
import requests
from dotenv import load_dotenv

load_dotenv(dotenv_path=os.path.join(os.getcwd(), ".env"))
URL = os.environ["SUPABASE_URL"]
KEY = os.environ["SUPABASE_SERVICE_KEY"]
XLSX = os.environ.get("LICENSES_XLSX", os.path.expanduser(
    "~/Downloads/2026-05-28-licenses-module.xlsx"))

H_WRITE = {
    "apikey": KEY, "Authorization": f"Bearer {KEY}",
    "Content-Profile": "beacon_v2", "Content-Type": "application/json",
    "Prefer": "resolution=merge-duplicates,return=minimal",
}


def clean_text(v):
    if v is None:
        return None
    s = str(v).replace("_x000D_", "").replace("\r", "\n").strip()
    return s or None


def as_str(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    return s or None


def as_date(v):
    if v is None:
        return None
    if isinstance(v, (dt.datetime, dt.date)):
        return v.date().isoformat() if isinstance(v, dt.datetime) else v.isoformat()
    s = str(v).strip()
    return s or None


def split_emails(v):
    if v is None:
        return []
    return [e.strip() for e in str(v).split(",") if e.strip()]


def main():
    apply = "--apply" in sys.argv
    ws = openpyxl.load_workbook(XLSX, data_only=True)["Sheet1"]

    rows = []
    for r in range(2, ws.max_row + 1):
        cell = [ws.cell(row=r, column=c).value for c in range(1, 12)]
        if cell[0] is None and cell[1] is None:
            continue
        enabled = cell[9]
        rows.append({
            "legacy_id":        int(cell[0]) if cell[0] is not None else None,
            "entity":           clean_text(cell[1]) or "(unnamed)",
            "state":            clean_text(cell[2]),
            "lic_type":         clean_text(cell[3]),
            "license_no":       as_str(cell[4]),
            "asce_m_no":        as_str(cell[5]),
            "first_issue_date": as_date(cell[6]),
            "expiration_date":  as_date(cell[7]),
            "notify_emails":    split_emails(cell[8]),
            "email_enabled":    (int(enabled) == 1) if enabled is not None else True,
            "notes":            clean_text(cell[10]),
        })

    print(f"Parsed {len(rows)} license rows from {XLSX}")
    print(f"{'entity':40s} {'state':5s} {'type':16s} {'expires':11s}  enabled  #emails")
    for x in rows:
        print(f"{(x['entity'] or '')[:40]:40s} {(x['state'] or '-'):5s} "
              f"{(x['lic_type'] or '-')[:16]:16s} {(x['expiration_date'] or '—'):11s}  "
              f"{'yes' if x['email_enabled'] else 'no ':7s}  {len(x['notify_emails'])}")

    if not apply:
        print("\nDry-run only. Re-run with --apply to upsert into beacon_v2.licenses.")
        return

    resp = requests.post(
        f"{URL}/rest/v1/licenses?on_conflict=legacy_id",
        headers=H_WRITE, json=rows, timeout=90,
    )
    if resp.status_code >= 400:
        print(f"[FAIL] status={resp.status_code} body={resp.text[:400]}")
        sys.exit(1)
    print(f"\nUpserted {len(rows)} license row(s) into beacon_v2.licenses.")


if __name__ == "__main__":
    main()
