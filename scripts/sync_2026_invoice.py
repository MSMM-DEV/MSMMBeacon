#!/usr/bin/env python3
"""
Wipe all year=2026 anticipated_invoice rows and reinsert them from the
`2026` worksheet of NEW_2026.xlsx so the Supabase table matches the
spreadsheet exactly.

Layout notes (observed in NEW_2026.xlsx, sheet "2026"):
  • Row 4  = header (Project No., Project Name, PM, Contract, …).
  • Row 5  = "ENG Projects" section marker.
  • Rows 6..37  = regular ENG projects.
  • Rows 38..44 = Orange projects (cell fill FFFFC000 on column C/Name).
  • Row 45 "TOTAL w/out ORG", Row 46 "TOTAL w/ ORG", Row 47 "* Not Included…"
  • Rows 51..53 = extra (on-hold) projects kept after the totals. They
    DO represent real invoice rows — this script includes them as
    non-Orange entries since they aren't amber-filled.
  • Rows 62+ = stray calc cells, ignored.

Column map (1-indexed):
  B=2  Project No.
  C=3  Project Name
  D=4  PM (single name, resolved to a user_id via roster match)
  E=5  Contract Amount
  N=14 MSMM Remaining to Bill as of 1/1/26
  O=15..S=19   Jan..May
  T=20         May Billing Notes (SKIP)
  U=21..AA=27  Jun..Dec

Orange handling: an Orange invoice row requires `source_potential_id`
pointing to a Potential where `probability='Orange'` (the UI's orange
tint + bottom-of-list ordering keys off that link). For each orange row
this script finds a matching Orange Potential (by project_number, else
project_name) or creates one.

This is destructive for year=2026 invoice rows (and their *_pms). Use
--dry-run to preview without writing.
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()
SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_SERVICE_KEY"]

REPO = Path(__file__).resolve().parent.parent
XLSX = REPO / "Data" / "Invoice Cycle Data" / "NEW_2026.xlsx"

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Accept-Profile": "beacon",
    "Content-Profile": "beacon",
    "Prefer": "return=representation",
}

ORANGE_RGB = "FFFFC000"  # cell fill on Orange invoice rows in the xlsx
MONTH_COLS = [
    "jan_amount", "feb_amount", "mar_amount", "apr_amount",
    "may_amount", "jun_amount", "jul_amount", "aug_amount",
    "sep_amount", "oct_amount", "nov_amount", "dec_amount",
]

# PM name → roster match hint. Two Scotts on staff: Scott Chehardy
# (short_name "Scott C.") and Scott Douglas ("Scott D."). The sheet
# writes "Scott" (ambiguous — maps to Douglas per backfill_pms.py) and
# "Scott C." (maps directly via short_name). No override needed for
# "Scott C." since the roster's short_name already matches.
PM_OVERRIDES = {
    "Scott": "Scott Douglas",
}


def api(method: str, path: str, **kw):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    r = requests.request(method, url, headers=HEADERS, **kw)
    if not r.ok:
        raise RuntimeError(f"{method} {path} → {r.status_code}: {r.text}")
    return r.json() if r.text else []


def num(v):
    """Coerce a cell value to a float for direct numeric columns. Empty → 0."""
    if v is None or v == "":
        return 0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0


def override_num(v):
    """Coerce a cell value for an OVERRIDE column. Empty stays NULL (auto-
    calc); numeric (including 0 and negatives) stays as the float — the
    spreadsheet's cached value is the source of truth."""
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def fill_rgb(cell):
    f = cell.fill
    if not f or not f.fgColor:
        return None
    try:
        rgb = f.fgColor.rgb
        if isinstance(rgb, str) and rgb != "00000000":
            return rgb
    except Exception:
        pass
    return None


def parse_worksheet() -> list[dict]:
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["2026"]
    out: list[dict] = []
    for i in range(6, ws.max_row + 1):
        proj_no = ws.cell(row=i, column=2).value
        name = ws.cell(row=i, column=3).value
        pm = ws.cell(row=i, column=4).value
        contract = ws.cell(row=i, column=5).value
        msmm_rem = ws.cell(row=i, column=14).value
        months = [ws.cell(row=i, column=c).value for c in (15, 16, 17, 18, 19, 21, 22, 23, 24, 25, 26, 27)]
        fill = fill_rgb(ws.cell(row=i, column=3))

        if not name or not isinstance(name, str):
            continue
        up = name.upper()
        if up in ("ENG PROJECTS", "PM PROJECTS"):
            continue
        if up.startswith("TOTAL"):
            continue
        if name.startswith("*") or "Not Included" in name:
            continue

        # Placeholder project numbers like 2026XX → NULL so the DB's
        # anticipated_invoice.project_number stays clean.
        if isinstance(proj_no, str) and re.fullmatch(r"\d{4}[Xx]+", proj_no):
            proj_no = None
        elif proj_no is not None:
            proj_no = str(proj_no).strip() or None

        # Pull both total columns (AB=28, AC=29) as OVERRIDES so the UI
        # shows exactly what the spreadsheet shows — critical because the
        # app's default Rollforward formula clamps negatives to 0 with
        # Math.max(0, …), which silently flipped three rows with legitimate
        # negative Rollforward values back to 0 on the last sync. YTD
        # Actual also persisted as an override to guarantee exact match
        # against any formula drift between sheet and app.
        ytd_override = override_num(ws.cell(row=i, column=28).value)
        rf_override  = override_num(ws.cell(row=i, column=29).value)

        out.append({
            "row": i,
            "project_number": proj_no,
            "project_name": name.strip(),
            "pm": (pm or "").strip() if isinstance(pm, str) else (str(pm) if pm else ""),
            "contract_amount": num(contract),
            "msmm_remaining_to_bill_year_start": num(msmm_rem),
            "months": [num(m) for m in months],
            "ytd_actual_override": ytd_override,
            "rollforward_override": rf_override,
            "is_orange": fill == ORANGE_RGB,
        })
    return out


def resolve_pm(name: str, users: list[dict]) -> str | None:
    if not name:
        return None
    probe = PM_OVERRIDES.get(name, name)
    parts = probe.split()
    for u in users:
        short = (u.get("short_name") or "").lower()
        first = (u.get("first_name") or "").lower()
        last = (u.get("last_name") or "").lower()
        disp = (u.get("display_name") or "").lower()
        if short and short == probe.lower():
            return u["id"]
        if disp == probe.lower():
            return u["id"]
        if len(parts) >= 2 and first == parts[0].lower() and last.startswith(parts[-1].lower()):
            return u["id"]
        if len(parts) == 1 and first == parts[0].lower():
            # Single-name match — accept only if unambiguous.
            candidates = [x for x in users if (x.get("first_name") or "").lower() == parts[0].lower()]
            if len(candidates) == 1:
                return u["id"]
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="Parse + plan without writing to Supabase.")
    args = ap.parse_args()

    print(f"Loading {XLSX.name}…")
    rows = parse_worksheet()
    reg = [r for r in rows if not r["is_orange"]]
    ora = [r for r in rows if r["is_orange"]]
    print(f"  parsed {len(rows)} rows ({len(reg)} regular, {len(ora)} orange)")

    print("\nFetching user roster for PM resolution…")
    users = api("GET", "users?select=id,first_name,last_name,display_name,short_name")
    print(f"  {len(users)} users")

    # Resolve PMs up-front so we can flag misses during dry-run.
    pm_map: dict[str, str | None] = {}
    for r in rows:
        if r["pm"] and r["pm"] not in pm_map:
            pm_map[r["pm"]] = resolve_pm(r["pm"], users)
    missing = [k for k, v in pm_map.items() if v is None and k]
    if missing:
        print(f"  ! PM names not matched to roster: {missing}")

    print("\nPlanning Orange Potential links…")
    existing_oranges = api(
        "GET",
        "potential_projects?probability=eq.Orange&select=id,project_number,project_name,year",
    )
    by_num = {p["project_number"]: p for p in existing_oranges if p.get("project_number")}
    by_name = {p["project_name"]: p for p in existing_oranges}

    orange_matched: list[tuple[dict, dict]] = []
    orange_to_create: list[dict] = []
    for r in ora:
        match = None
        if r["project_number"] and r["project_number"] in by_num:
            match = by_num[r["project_number"]]
        elif r["project_name"] in by_name:
            match = by_name[r["project_name"]]
        if match:
            orange_matched.append((r, match))
        else:
            orange_to_create.append(r)
    print(f"  {len(orange_matched)} orange rows match existing Orange potentials")
    print(f"  {len(orange_to_create)} orange rows need a new Orange potential")

    print("\nExisting 2026 invoices in DB:")
    existing_inv = api("GET", "anticipated_invoice?year=eq.2026&select=id")
    print(f"  {len(existing_inv)} rows (will be replaced)")

    if args.dry_run:
        print("\n--dry-run: no writes performed.")
        return

    # --- Create any missing Orange potentials ---
    orange_potential_ids: dict[str, str] = {r["project_name"]: m["id"] for r, m in orange_matched}
    if orange_to_create:
        payload = [
            {
                "year": 2026,
                "project_name": r["project_name"],
                "project_number": r["project_number"],
                "role": "Prime",
                "probability": "Orange",
                "total_contract_amount": r["contract_amount"],
            }
            for r in orange_to_create
        ]
        created = api("POST", "potential_projects", json=payload)
        for r, c in zip(orange_to_create, created):
            orange_potential_ids[r["project_name"]] = c["id"]
            print(f"  + created Orange potential · {r['project_name'][:50]}")

    # --- Wipe existing 2026 invoices (cascades anticipated_invoice_pms) ---
    if existing_inv:
        print("\nDeleting existing 2026 invoice rows…")
        api("DELETE", "anticipated_invoice?year=eq.2026")

    # --- Insert fresh invoice rows ---
    print("Inserting fresh invoice rows…")
    # PostgREST requires uniform keys across a batch insert. Always include
    # source_potential_id + both overrides (null when not applicable) so
    # every object in the array shares the same shape.
    payloads = []
    for r in rows:
        p = {
            "year": 2026,
            "project_name": r["project_name"],
            "project_number": r["project_number"],
            "contract_amount": r["contract_amount"],
            "type": "ENG",
            "msmm_remaining_to_bill_year_start": r["msmm_remaining_to_bill_year_start"],
            "source_potential_id": orange_potential_ids[r["project_name"]] if r["is_orange"] else None,
            "ytd_actual_override": r["ytd_actual_override"],
            "rollforward_override": r["rollforward_override"],
        }
        for col, v in zip(MONTH_COLS, r["months"]):
            p[col] = v
        payloads.append(p)

    inserted = api("POST", "anticipated_invoice", json=payloads)
    print(f"  inserted {len(inserted)}")

    # --- Insert PMs ---
    pm_rows = []
    for inv_row, src in zip(inserted, rows):
        uid = pm_map.get(src["pm"])
        if uid:
            pm_rows.append({"anticipated_invoice_id": inv_row["id"], "user_id": uid})
    if pm_rows:
        api("POST", "anticipated_invoice_pms", json=pm_rows)
        print(f"  inserted {len(pm_rows)} PM tags")

    print("\n✓ Sync complete.")


if __name__ == "__main__":
    main()
